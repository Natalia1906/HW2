#TODO задание 2
from openpyxl import Workbook, load_workbook
import re
def extra():
    col1 = []
    workbook = load_workbook('Лист1.xlsx')
    sheet = workbook.active
    for i in range(1, sheet.max_row + 1):
        col1.append(sheet.cell(row=i, column=1).value)

    col2 = []
    workbook = load_workbook('Лист2.xlsx')
    sheet = workbook.active
    for i in range(1, sheet.max_row + 1):
        col2.append(sheet.cell(row=i, column=2).value)

    col3 = []
    workbook = load_workbook('Лист3.xlsx')
    sheet = workbook.active
    for i in range(1, sheet.max_row + 1):
        col3.append(sheet.cell(row=i, column=3).value)

    matrix = [col1, col2, col3]

    combined_file = Workbook()
    combined_sheet = combined_file.active
    row_i = 0
    for i in matrix:
        row_i += 1
        col_i = 0
        for j in i:
            col_i += 1
            combined_sheet.cell(row=row_i, column=col_i, value=j)
    combined_file.save('combined_file.xlsx')

#TODO Домашнее задание №26. Регулярные выражения
#1
emails_list = ["admin@gmail.com", "user@email.com", "admin123@gmail.com", "user34@email.com"]
def extract_domains(emails):
    domains = []
    pattern = r'@([\w.-]+)'

    for email in emails:
        match = re.search(pattern, email)
        if match:
            domain = match.group()
            domains.append(domain)
    return domains

#2
def words(text):
    pattern1 = r'\b\w*[aeiouAEIOU]\w*\b'
    need_word = re.findall(pattern1, text)
    return need_word

#3
def string(text2):
    pattern2 = r'[,;\s]'
    splitted = re.split(pattern2, text2)
    return splitted

#TODO Домашнее задание №27. Модули и пакеты

candidates = ["Аскаров", "Бекмуханов", "Ернур", "Пешая", "Карим", "Шаримазданов"]

def search_winner(candidates, votes):
    vote_counts = {candidate: votes.count(candidate) for candidate in candidates}
    max_votes = max(vote_counts.values())
    winners = [candidate for candidate, count in vote_counts.items() if count == max_votes]
    winner = min(winners,key = len)
    return winner, max_votes

votes = []
while True:
    vote = input("Вы отдаете голос за(Enter для завершения голосования): ")
    if vote == "":
        break
    votes.append(vote)

if __name__ == "__main__":
    extra()
    print(extract_domains(emails_list))
    print(words("Apple is a fruit"))
    print(string("apple;banana,cherry,orange"))
    print(search_winner(candidates, votes))
